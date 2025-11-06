import os
import cv2
import numpy as np
from pyzbar.pyzbar import decode
import time
from openpyxl import load_workbook

authorized_users = np.array([])
# read file
wb = load_workbook(filename='./students.xlsx')
ws = wb['students']
label = []
students_list = np.array([])
for col in list(ws.iter_cols()):
    label.append(col[0].value)
# print(label)

for row in list(ws.iter_rows(values_only=True))[1:]:
    row_dict = {}
    # print(row)
    for i in range(len(row)):
        row_dict[label[i]] = row[i]
    students_list = np.append(students_list, row_dict)

# print(students_list)

# print('----------------------')
# print(students_list)
# print(students_list.shape)
# print(students_list.shape)
cap = cv2.VideoCapture(0)
while (True):
    font = cv2.FONT_HERSHEY_SIMPLEX
    font_color = (255, 255, 255)
    font_scale = 1
    font_thicknes = 2
    # prior to time to reading camera
    start_time = time.time()
    ret, frame = cap.read()
    end_time = time.time()
    qr_info = decode(frame)
    # print(frame.shape)
    # caculate number of frames per second
    fps = 1 / (end_time - start_time)
    cv2.putText(frame, "FPS: {:.2f}".format(fps), (0, 30), font, font_scale, font_color, font_thicknes)

    if len(qr_info) > 0:
        check_student_value = 0
        qr = qr_info[0]
        data_student_code = qr.data.decode('utf-8')
        rect = qr.rect
        # print('Student code:', data_student_code)
        # print('Student code type:', type(data_student_code))
        polygon = qr.polygon
        print(students_list)
        for i in range(len(students_list)):
            # print('students_list[i]: ', students_list[i]['StudentCode'])
            # print('students_list[i] type: ', type(students_list[i]['StudentCode']))
        
            if (data_student_code.strip() == students_list[i][
                'StudentCode'].strip()) and (students_list[i]
                                             ['diemDanh'].strip() == 'unchecked'):
                ws[f"D{i + 2}"] = 'checked'  # pass index inlist and excel so have to plus 2
                cv2.putText(frame, 'Roll call successful', (rect.left, rect.top - 20), cv2.FONT_HERSHEY_COMPLEX, 1,
                            (0, 255, 0), 2)
                wb.save("students.xlsx")
                # print('01')
                check_student_value = 1
            elif (data_student_code.strip() == students_list[i][
                'StudentCode'].strip()) and (students_list[i]
                                             ['diemDanh'].strip() == 'checked'):
                cv2.putText(frame, 'Presented', (rect.left, rect.top - 20), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 2)
                # print('02')
                check_student_value = 1
        if check_student_value == 0:
            cv2.putText(frame, 'Invalid Student', (rect.left, rect.top),
                        cv2.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 2)
            # print('03')
        frame = cv2.rectangle(frame, (rect.left, rect.top), (rect.left + rect.width, rect.top + rect.height),
                              (0, 255, 0), 5)
        frame = cv2.polylines(frame, [np.array(polygon)], True, (255, 0, 0), 5)
    cv2.imshow("Roll up", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

wb.close()
cap.release()
cv2.destroyAllWindows()
